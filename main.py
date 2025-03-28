from flask import Flask, render_template, send_from_directory, request, redirect, url_for, send_file
from openpyxl import load_workbook, Workbook
import matplotlib.pyplot as plt
import os
import urllib  # Added missing import

app = Flask(__name__)

EXCEL_FILE = os.path.join(os.getcwd(), "info_kopsavilkums.xlsx")  # Ensure platform-independent paths
BASE_DIR = os.getcwd()
COMMENTS_FILE = os.path.join(BASE_DIR, "comments.txt")
CHART_FILE = os.path.join(BASE_DIR, "biotops_ekologija.png")
DEFAULT_IMAGE = os.path.join(BASE_DIR, "default.jpg")  # Ensure a default image file exists

def read_excel():
    try:
        if not os.path.exists(EXCEL_FILE):
            return [{"Kļūda": "Excel fails netika atrasts!"}]
        
        wb = load_workbook(EXCEL_FILE)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1] if cell.value]  # Safeguard for missing headers
        cats = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            cat_info = dict(zip(headers, row))
            cat_name = cat_info.get("Nosaukums", "").strip()
            image_filename = f"{cat_name}.jpg"
            image_path = os.path.join(BASE_DIR, image_filename)

            cat_info["Attēls"] = f"/image/{image_filename}" if os.path.exists(image_path) else f"/image/{DEFAULT_IMAGE}"

            cats.append(cat_info)

        return cats
    except Exception as e:
        return [{"Kļūda": f"Neizdevās atvērt Excel failu: {e}"}]

@app.route('/')
def home():
    order = request.args.get("order", "asc")
    cats = read_excel()
    cats = sorted(cats, key=lambda x: x.get("Nosaukums", ""), reverse=(order == "desc"))  # Use get to avoid KeyError
    return render_template('index.html', cats=cats, order=order)

@app.route('/cat/<path:name>')
def cat_detail(name):
    cats = read_excel()
    decoded_name = urllib.parse.unquote(name).replace('-', ' ')
    
    selected_cat = next((cat for cat in cats if cat['Nosaukums'].strip().lower() == decoded_name.strip().lower()), None)

    if selected_cat:
        return render_template('cat.html', cat=selected_cat)
    else:
        return "Kaķis nav atrasts!", 404

@app.route('/image/<filename>')
def get_image(filename):
    return send_from_directory(BASE_DIR, filename)

@app.route('/comments', methods=["GET", "POST"])
def comments():
    if request.method == "POST":
        comment = request.form.get("comment")
        if comment:
            with open(COMMENTS_FILE, "a", encoding="utf-8") as file:
                file.write(comment + "\n")
        return redirect(url_for('comments'))

    if os.path.exists(COMMENTS_FILE):
        with open(COMMENTS_FILE, "r", encoding="utf-8") as file:
            comments = file.readlines()
    else:
        comments = []

    return render_template('comments.html', comments=comments)

@app.route('/delete_comment/<int:comment_index>', methods=["POST"])
def delete_comment(comment_index):
    if os.path.exists(COMMENTS_FILE):
        with open(COMMENTS_FILE, "r", encoding="utf-8") as file:
            comments = file.readlines()

        if 0 <= comment_index < len(comments):
            del comments[comment_index]

            with open(COMMENTS_FILE, "w", encoding="utf-8") as file:
                file.writelines(comments)

    return redirect(url_for('comments'))

@app.route('/download')
def download_excel():
    try:
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Kaķi"

        cats = read_excel()
        
        if cats and isinstance(cats, list) and isinstance(cats[0], dict):
            headers = cats[0].keys()
            sheet.append(list(headers))

            for cat in cats:
                sheet.append([cat.get(key, "") for key in headers])
        else:
            return "Nav derīgu datu, ko eksportēt!", 400

        download_path = os.path.join(BASE_DIR, "info_kopsavilkums.xlsx")
        wb.save(download_path)

        return send_file(download_path, as_attachment=True, download_name="info_kopsavilkums.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        return f"Kļūda ģenerējot failu: {e}", 500

@app.route('/chart')
def generate_chart():
    try:
        cats = read_excel()
        biotops_count = {}
        for cat in cats:
            biotops = cat.get("Biotops un ekoloģija", "Nepieejams")
            biotops_count[str(biotops)] = biotops_count.get(str(biotops), 0) + 1

        plt.figure(figsize=(10, 6))
        plt.bar(biotops_count.keys(), biotops_count.values(), color='skyblue')
        plt.title("Biotops un ekoloģija")
        plt.xlabel("Biotops")
        plt.ylabel("Skaits")
        plt.xticks(rotation=45)
        plt.tight_layout()

        chart_path = os.path.join(BASE_DIR, CHART_FILE)
        plt.savefig(chart_path)

        return send_file(chart_path, mimetype='image/png', as_attachment=True, download_name=CHART_FILE)
    except Exception as e:
        return f"Kļūda veidojot diagrammu: {e}", 500

if __name__ == '__main__':
    app.run(debug=True)
